"""/   PROJECT DESCRIPTION """

"""/ Face recognition project developed using Opencv + Python bindings.The project uses "face_recognition" dlib library trained 
     on deep learning.Frames are streamed directly from the raspberry pi(client) to the server using "imagezmq" python library 
     which is based on 'zmq' protocol.An encoding file(encoding.pickle) which contains the encodings of the faces of persons is
     stored before hand.Frame processing that is encoding each incoming frame and comparing that encoding with the already 
     stored encoding of the faces takes place on the server side.Once the face is recognized a message "DETECTED" is sent to the
     raspberry pi.
     Based on this message, LED attached to the raspberry pi changes its color which shows the confirmation to the user that 
     the face is recognized.The frames and message are exchanged between the raspberry pi and the server using the local 
     network IP.Mutiple raspberry pi's can send their live feed at the same time to the server for face recognition process.
     An electronic-mail is also sent at the end of the day with the attachment containing the attendance details of that day.
"""
"""/     SERVER CODE """

import cv2 as cv
import imagezmq
from imutils import build_montages
import face_recognition
import openpyxl as op
import numpy
import datetime
import imutils
import pickle
import argparse
import os
import os.path
import socket
import zmq
import time
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import Workbook


TCP_IP = '192.168.86.57'
TCP_PORT = 6000
Message = 'DETECTED'   
exitMessage = 'EXIT'


ap = argparse.ArgumentParser()
ap.add_argument("-e","--encodings",required=True,help="path to serialized db of facial encodings")
ap.add_argument("-o","--output",type=str,help="path to output video")
ap.add_argument("-y","--display",type=int,default=0,help="wheather or not to display output frame to screen")
ap.add_argument("-d","--detection-method",type=str,default="hog",help="face detection model to use: either hog or cnn")
ap.add_argument("-mW","--montageW",required=True,type=int,help="montage frame width")
ap.add_argument("-mH","--montageH",required=True,type=int,help="montage frame height")
args = vars(ap.parse_args()) 


imageHub = imagezmq.ImageHub()


print("[INFO] loading encodings...")
pickle.loads(open(args["encodings"], "rb").read())

lastActive = {}
lastActiveCheck = datetime.datetime.now()


mW = args["montageW"]
mH = args["montageH"]

name = "Unknown"
names = []
face_locations = []
face_encodings = []
framedict = {}
totalframes = 1
totaldays = 0
m = 2
flag = 0
temp = 1



now= datetime.datetime.now()
today=now.day
month=now.month

value = os.path.isfile('/home/yatharth/project/'+ str(today)+'.xlsx')
if bool(value) == True:
    pass
else:
    wb = Workbook()
    wb.save(filename = str(today)+'.xlsx')

while True:
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    s.connect((TCP_IP, TCP_PORT))
    totalframes = totalframes + 1
    (rpiName , frame) = imageHub.recv_image()              
    imageHub.send_reply(b'OK')
    cv.imshow("Livefeed",frame)
    rgb_frame = frame[: , : , ::-1]                  

    face_locations = face_recognition.face_locations(rgb_frame , model=args["detection_method"])
    face_encodings = face_recognition.face_encodings(rgb_frame,face_locations)      


    for face_encoding in face_encodings:
        data = pickle.loads(open(args["encodings"], "rb").read())            
        matches = face_recognition.compare_faces(data["encodings"],face_encoding,0.4) 



        if True in matches:
            matchedIds = [i for (i,b) in enumerate(matches) if b]
            for i in matchedIds:
                name = data["names"][i]    
                print(name)
                s.send(Message.encode())     
                data = s.recv(1024).decode() 
                
              

                book = op.load_workbook("INFORMATION.xlsx")
                book1 = op.load_workbook(str(today)+'.xlsx')
                book2 = op.load_workbook("presentdetails.xlsx")


                sheet = book.active          #INFORMATION SHEET
                sheet1 = book1.active        #IN-TIME , OUT-TIME SHEET
                sheet2 = book2[str(month)]   #PRESENT-DAYS COUNT SHEET

                for j in range(2,40):
                    if sheet.cell( row = j ,column = 2 ).value == name:
                        while sheet1.cell(row = m , column = 2).value != None:
                            m = m + 1
                        sheet1.cell(row = m , column = 2).value = name
                        sheet1.cell(row = m , column = 1).value = sheet.cell(row = j , column = 1).value
                        sheet1.cell(row = m , column = 3).value = sheet.cell(row = j , column = 3).value
                        sheet1.cell(row = m , column = 4).value = datetime.datetime.now()
                        m = m + 0

                        sheet2.cell( row = j , column = int(today) + 4 ).value = "PRESENT"
                        for k in range (4,40):
                            if sheet2.cell(row = j , column = k).value == "PRESENT":
                                totaldays = totaldays + 1
                        sheet2.cell(row = j , column = 3).value = totaldays 
                        break

                book1.save(str(today)+'.xlsx') 
                book2.save("presentdetails.xlsx")


    t=time.localtime()

    current_time = time.strftime("%H:%M:%S",t)
    if (current_time >= '17:30:00' and current_time <= '17:30:10')and temp == 1:
        fromaddr = "yatharth@iimjobs.com"
        toaddr = "yatharth@iimjobs.com"

        msg = MIMEMultipart()
        msg['From'] = fromaddr
        msg['To'] = toaddr
        body = "Today's attendance details"
        msg.attach(MIMEText(body,'plain'))

        filename = str(today)+'.xlsx'
        attachment = open("/home/yatharth/project/"+str(today)+'.xlsx',"rb")

        p =MIMEBase('application','octet-stream')

        p.set_payload((attachment).read())
        encoders.encode_base64(p)
        p.add_header('Content-Disposition',"attachment;filename= %s" %filename)
        msg.attach(p)
        s=smtplib.SMTP('smtp.gmail.com',587)
        s.starttls()
        s.login(fromaddr,"redyellow")
        text = msg.as_string()
        s.sendmail(fromaddr,toaddr,text)
        temp = 0
        print("Mail sent")
        s.quit()
    totaldays = 0
    name = "Unknown"
    flag = 0

    key = cv.waitKey(1)
    if  key == ord('q'):
        s.send(exitMessage.encode())        
        data = s.recv(20).decode()     
        break
